import subprocess
import openpyxl
from openpyxl import Workbook, load_workbook
import datetime
import os
import time
from persiantools.jdatetime import JalaliDate
from openpyxl.styles import Font, Alignment, PatternFill

RESULTS_FILE = "test_AllTestResults.xlsx"
SCREENSHOTS_DIR = "screenshots"

target_date = "1404-03-02"
start_time = "11:29"
repeat_count = 1
interval_minutes = 0

os.makedirs(SCREENSHOTS_DIR, exist_ok=True)

test_name_mapping = {
    "pages/test_login.py": "تست ورود کاربر",
    "pages/test_signup.py": "تست ثبت نام در سامانه",
}

tests_in_order = list(test_name_mapping.keys())

def apply_font_b_nazanin(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell.font = Font(name='B Nazanin')

def adjust_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 5, 50)

def insert_test_run_header_to_excel():
    now = datetime.datetime.now()
    timestamp = now.strftime("%H:%M:%S")
    datestamp = JalaliDate(now).isoformat()
    title_text = f"نتیجه‌ی تست اتومیشن سامانه‌ی صورت‌حساب با Playwright - تاریخ: {datestamp} - ساعت شروع تست: {timestamp}"

    if not os.path.exists(RESULTS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"
    else:
        wb = load_workbook(RESULTS_FILE)
        ws = wb.active

    insert_row = ws.max_row + 1

    # عنوان کلی - ردیف merge شده، پس‌زمینه مشکی و فونت سفید وسط‌چین
    ws.insert_rows(insert_row)
    ws.merge_cells(start_row=insert_row, start_column=1, end_row=insert_row, end_column=6)
    title_cell = ws.cell(row=insert_row, column=1)
    title_cell.value = title_text
    title_cell.fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")
    title_cell.font = Font(name='B Nazanin', bold=True, size=13, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # هدر ستون‌ها
    headers = ["نام تست", "تاریخ", "زمان", "وضعیت", "فایل عکس", "تعداد تست کیس‌های هر سناریو"]
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_row = insert_row + 1

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = Font(name='B Nazanin', bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    apply_font_b_nazanin(ws)
    adjust_column_widths(ws)
    wb.save(RESULTS_FILE)

def count_test_cases(test_file):
    result = subprocess.run(["pytest", "--collect-only", "-q", test_file], capture_output=True, text=True)
    lines = result.stdout.strip().splitlines()
    test_lines = [line for line in lines if "::" in line and not line.startswith("<")]
    return len(test_lines), test_lines  # return also list for further comparison

def log_result_to_excel(test_file, status, screenshot_path="", num_tests=None):
    test_name = test_name_mapping.get(test_file, test_file)
    now = datetime.datetime.now()
    timestamp = now.strftime("%H:%M:%S")
    datestamp = JalaliDate(now).isoformat()
    status_text = "موفق" if status else "ناموفق"

    if not os.path.exists(RESULTS_FILE):
        wb = Workbook()
        ws = wb.active
    else:
        wb = load_workbook(RESULTS_FILE)
        ws = wb.active

    ws.append([test_name, datestamp, timestamp, status_text, screenshot_path or "", num_tests])
    row_index = ws.max_row

    status_color = "C6EFCE" if status else "FFC7CE"
    ws.cell(row=row_index, column=4).fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
    ws.cell(row=row_index, column=4).font = Font(name='B Nazanin', bold=True)
    ws.cell(row=row_index, column=4).alignment = Alignment(horizontal="center")

    for col in [1, 2, 3, 5, 6]:
        cell = ws.cell(row=row_index, column=col)
        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        cell.font = Font(name='B Nazanin')
        cell.alignment = Alignment(horizontal="center")

    apply_font_b_nazanin(ws)
    adjust_column_widths(ws)
    wb.save(RESULTS_FILE)


def run_tests():
    print("=" * 80)
    print("🎯 شروع اجرای تست‌ها...\n")

    insert_test_run_header_to_excel()

    for test_file in tests_in_order:
        print(f"🚀 در حال اجرای: {test_file}")
        total_cases, case_ids = count_test_cases(test_file)

        result = subprocess.run(
            ["pytest", "-v", "--tb=short", "--disable-warnings", test_file],
            capture_output=True, text=True
        )

        passed_cases = 0
        for line in result.stdout.splitlines():
            if "::" in line and "PASSED" in line:
                passed_cases += 1

        failed_cases = total_cases - passed_cases

        # ✅ تعیین وضعیت بر اساس مقایسه تعداد تست‌کیس‌های پاس شده و رد شده
        passed = passed_cases >= failed_cases

        log_result_to_excel(test_file, passed, "", total_cases)

        # ذخیره آمار برای محاسبه گزارش کلی
        global test_result_summary
        test_result_summary.append({
            "file": test_file,
            "total": total_cases,
            "passed": passed_cases,
            "failed": failed_cases,
            "status": passed
        })

    print("✅ اجرای تست‌ها پایان یافت.")
    print("=" * 80)






    print("✅ اجرای تست‌ها پایان یافت.")
    print("=" * 80)

def wait_until_start():
    year, month, day = map(int, target_date.split("-"))
    j_date = JalaliDate(year, month, day)
    t_hour, t_min = map(int, start_time.split(":"))
    start_dt = datetime.datetime.combine(j_date.to_gregorian(), datetime.time(t_hour, t_min))
    now = datetime.datetime.now()
    seconds = (start_dt - now).total_seconds()
    if seconds > 0:
        print(f"⏳ انتظار تا ساعت شروع {start_time} در تاریخ {target_date} ...")
        time.sleep(seconds)
    else:
        print("⚠️ زمان شروع گذشته! اجرای فوری...")

def calculate_success_rate(start_time):
    if not os.path.exists(RESULTS_FILE):
        print("❌ فایل گزارش موجود نیست.")
        return

    wb = load_workbook(RESULTS_FILE)
    ws = wb.active

    total_scenarios = len(test_result_summary)
    total_testcases = sum(item["total"] for item in test_result_summary)
    passed_testcases = sum(item["passed"] for item in test_result_summary)
    failed_testcases = sum(item["failed"] for item in test_result_summary)
    passed_scenarios = sum(1 for item in test_result_summary if item["status"])
    failed_scenarios = total_scenarios - passed_scenarios

    success_percent = round((passed_scenarios / total_scenarios) * 100, 2)
    fail_percent = round((failed_scenarios / total_scenarios) * 100, 2)

    now = datetime.datetime.now()
    duration = str(now - start_time).split('.')[0]
    end_time = now.strftime("%H:%M:%S")
    result_status = "وضعیت سامانه پایدار است" if success_percent >= 80 else \
                    "وضعیت سامانه نیمه‌پایدار است (ریسک بالا)" if 50 < success_percent < 80 else \
                    "وضعیت سامانه ناپایدار است"

    ws.append([
        "📊 نتیجه کلی تست",
        f"مدت زمان اجرا: {duration}",
        f"زمان پایان تست: {end_time}",
        f"تعداد کل سناریو تست‌ها: {total_scenarios}",
        f"تعداد کل تست‌کیس‌ها: {total_testcases}",
        f"تعداد کل تست‌کیس‌های پاس شده: {passed_testcases}"
    ])

    ws.append([
        "📋 وضعیت تست‌ها",
        f"{passed_scenarios} سناریو پاس شده ({success_percent}%)",
        f"{failed_scenarios} سناریو رد شده ({fail_percent}%)",
        result_status,
        "",
        f"تعداد کل تست‌کیس‌های رد شده: {failed_testcases}"
    ])

    for row_num in [ws.max_row - 1, ws.max_row]:
        for col in range(1, 7):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(name='B Nazanin', bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    apply_font_b_nazanin(ws)
    adjust_column_widths(ws)
    wb.save(RESULTS_FILE)

# اجرای کلی برنامه
test_result_summary = []

wait_until_start()
execution_start_time = datetime.datetime.now()

for i in range(repeat_count):
    print(f"\n🕐 اجرای نوبت {i+1} از {repeat_count}")
    run_tests()
    if i < repeat_count - 1:
        print(f"⌛ انتظار {interval_minutes} دقیقه برای اجرای بعدی...")
        time.sleep(interval_minutes * 60)

calculate_success_rate(execution_start_time)
