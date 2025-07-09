import re
import os
import datetime
import logging
from time import sleep
from pytest_bdd import scenario, when, then, parsers, given
from faker import Faker
from openpyxl import Workbook, load_workbook
from playwright.sync_api import expect
import pytest
from playwright.sync_api import sync_playwright
from functools import wraps

@pytest.fixture(scope="function")
def page():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)  # True برای تست‌های CI
        context = browser.new_context(
            viewport={"width": 1920, "height": 1080},  # نمایشگر بزرگ
            screen={"width": 1920, "height": 1080}
        )
        page = context.new_page()
        yield page
        context.close()
        browser.close()

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

RESULTS_FILE = "test_results.xlsx"
SCREENSHOT_DIR = "./screenshots"
os.makedirs(SCREENSHOT_DIR, exist_ok=True)


def log_result(test_name, success=True, screenshot_path=None):
    date_now = datetime.datetime.now().strftime("%Y-%m-%d")
    time_now = datetime.datetime.now().strftime("%H:%M:%S")
    result = "موفق" if success else "ناموفق"

    if not os.path.exists(RESULTS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["تست", "تاریخ", "ساعت", "نتیجه تست", "مسیر اسکرین‌شات"])
    else:
        wb = load_workbook(RESULTS_FILE)
        ws = wb.active

    ws.append([test_name, date_now, time_now, result, screenshot_path or "-"])
    wb.save(RESULTS_FILE)
    logger.info(f"نتیجه تست '{test_name}' با وضعیت: {result}")


def take_screenshot(page, test_name):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    screenshot_path = os.path.join(SCREENSHOT_DIR, f"{test_name}_{timestamp}.png")
    page.screenshot(path=screenshot_path)
    logger.warning(f"📸 اسکرین‌شات گرفته شد: {screenshot_path}")
    return screenshot_path

def log_step(test_name):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            page = kwargs.get("page", None)
            try:
                return func(*args, **kwargs)
            except Exception:
                if page:
                    screenshot_path = take_screenshot(page, test_name)
                else:
                    screenshot_path = None
                log_result(test_name, success=False, screenshot_path=screenshot_path)
                raise
        return wrapper
    return decorator


# ========== تست اول: ورود صحیح ==========

@pytest.mark.order(1)
@scenario("../features/login.feature", "Correct login")
def test_Login_Correct():
    pass


@given("کاربر در صفحه ورود به سامانه قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")


@when("کاربر اطلاعات ورود را به صورت صحیح وارد می‌کند")
def fill_Login_Correct_test(page):
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071523")
    sleep(0.5)
    page.get_by_placeholder(" رمز عبور را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder(" رمز عبور را وارد کنید").fill("0081071523")
    sleep(0.5)
    page.get_by_placeholder("کد امنیتی را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("کد امنیتی را وارد کنید").fill("PASSED")
    sleep(0.5)
    page.get_by_role("button", name="ورود").click()
    sleep(1)
    page.fill("#otpCode", "12345")
    sleep(1)
    page.get_by_role("button", name="ورود").click()
    sleep(1)


@then("کاربر با موفقیت وارد سامانه می‌شود و داشبورد نمایش داده می‌شود")
def check_dashboard_loaded(page):
    # صبر کن المان داشبورد ظاهر بشه
    try:
        expect(page.locator("text=برگزاری جلسات مجامع")).to_be_visible(timeout=15000)
        print("✅ ورود موفق و داشبورد نمایش داده شد.")
    except:
        print("❌ ورود موفق نبود یا داشبورد ظاهر نشد.")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        raise


@then(parsers.parse("کاربر با موفقیت وارد سامانه می‌شود و داشبورد نمایش داده می‌شود"))
@log_step("تست ورود صحیح و نمایش داشبورد")
def check_dashboard_loaded(page):
    expect(page.locator("text='برگزاری جلسات مجامع'")).to_be_visible()
    log_result("تست ورود صحیح و نمایش داشبورد", success=True)


# ========== تست دوم: نام کاربری اشتباه ==========

@pytest.mark.order(2)
@scenario("../features/login.feature", "Login with incorrect username")
def test_Login_incorrect_username():
    pass


@given("کاربر در صفحه ورود به سامانه قرار دارد")
def open_login_page_username_error(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")


@when("کاربر نام کاربری  نامعتبر را وارد می‌کند")
def fill_invalid_username(page):
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071000")
    sleep(0.5)


@when("کاربر رمز عبور و کد امنیتی معتبر را وارد می‌کند")
def fill_correct_password_and_captcha(page):
    sleep(0.5)
    page.get_by_placeholder(" رمز عبور را وارد کنید").fill("0081071523")
    sleep(0.5)
    page.get_by_placeholder("کد امنیتی را وارد کنید").fill("PASSED")
    sleep(0.5)


@when("کاربر روی دکمه ورود کلیک می‌کند")
def click_login_button_wrong_user(page):

    sleep(0.5)
    page.get_by_role("button", name="ورود").click()
    sleep(1)


@then(parsers.parse("پیغام خطای ورود با نام کاربری اشتباه نمایش داده می‌شود"))

@log_step("تست ورود ناموفق با نام کاربری اشتباه")

def check_invalid_login_error(page):
    sleep(2)
    expect(page.locator("text='خطا در ورود به سامانه'")).to_be_visible()
    log_result("تست عدم ورود با نام کاربری اشتباه", success=True)


# ========== تست سوم: پسورد اشتباه ==========

@pytest.mark.order(3)
@scenario("../features/login.feature", "Login with incorrect password")
def test_Login_incorrect_Password():
    pass


@given("کاربر در صفحه ورود به سامانه قرار دارد")
def open_login_page_password_error(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")


@when("کاربر رمز عبور نامعتبر را وارد می‌کند")
def fill_invalid_Password(page):
    sleep(0.5)
    page.get_by_placeholder(" رمز عبور را وارد کنید").fill("00810715211")
    
    sleep(0.5)


@when("کاربر نام کاربری و کد امنیتی صحیح را وارد می‌کند")
def fill_correct_username_and_captcha(page):
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071000")
    sleep(0.5)
    page.get_by_placeholder("کد امنیتی را وارد کنید").fill("PASSED")
    sleep(0.5)


@when("کاربر روی دکمه ورود کلیک می‌کند")
def click_login_button_wrong_user(page):

    sleep(0.5)
    page.get_by_role("button", name="ورود").click()
    sleep(1)


@then(parsers.parse("پیغام خطای ورود با رمز عبور اشتباه نمایش داده می‌شود"))

@log_step("تست ورود ناموفق با پسورد اشتباه")

def check_invalid_password_error(page):
    sleep(2)
    expect(page.locator("text='خطا در ورود به سامانه'")).to_be_visible()
    log_result("تست عدم ورود با پسورد اشتباه", success=True)


# ========== تست چهارم: کد کپچا ی اشتباه ==========


@pytest.mark.order(4)
@scenario("../features/login.feature", "Login with incorrect captcha code")
def test_Login_incorrect_captcha():
    pass


@given("کاربر در صفحه ورود به سامانه قرار دارد")
def open_login_page_captcha_error(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")


@when("کاربر نام کاربری و رمز عبور صحیح را وارد می‌کند")
def fill_valid_username_Password(page):
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071523")
    sleep(0.5)
    page.get_by_placeholder(" رمز عبور را وارد کنید").fill("0081071523")
    
    sleep(0.5)


@when("کاربر کد کپچای نادرست را وارد می‌کند")
def fill_invalid_captcha(page):
    sleep(0.5)
    page.get_by_placeholder("کد امنیتی را وارد کنید").fill("invalid")
    sleep(0.5)


@when("کاربر روی دکمه ورود کلیک می‌کند")
def click_login_button_wrong_user(page):

    sleep(0.5)
    page.get_by_role("button", name="ورود").click()
    sleep(3)


@then(parsers.parse("پیغام خطای ورود با کد کپچا اشتباه نمایش داده می‌شود"))

@log_step("تست ورود ناموفق با کد کپچا اشتباه")

def check_invalid_captcha_error(page):
    sleep(2)
    expect(page.locator("text='خطا در ورود به سامانه'")).to_be_visible()
    log_result("تست عدم ورود با کد کپچا اشتباه", success=True)

# ========== تست پنجم: عدم درج نام کاربری ==========

@pytest.mark.order(5)
@scenario("../features/login.feature", "Login without username")
def test_Login_without_username():
    pass


@given("کاربر در صفحه ورود به سامانه قرار دارد")
def open_login_page_without_username(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")


@when("کاربر نام کاربری را وارد نمی‌کند")
def fill_without_username(page):
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").fill("")
    sleep(0.5)


@when("کاربر رمز عبور و کد امنیتی معتبر را وارد می‌کند")
def fill_correct_password1_and_captcha1(page):
    sleep(0.5)
    page.get_by_placeholder(" رمز عبور را وارد کنید").fill("0081071523")
    sleep(0.5)
    page.get_by_placeholder("کد امنیتی را وارد کنید").fill("PASSED")
    sleep(0.5)


@when("کاربر روی دکمه ورود کلیک می‌کند")
def click_login_button_without_user(page):

    sleep(0.5)
    page.get_by_role("button", name="ورود").click()
    sleep(1)


@then(parsers.parse("پیغام خطای ورود با عدم درج نام کاربری نمایش داده می‌شود"))

@log_step("تست ورود ناموفق با نام کاربری اشتباه")

def check_invalid_login_error_without_username(page):
    sleep(2)
    expect(page.locator("text='فیلد الزامی است.'")).to_be_visible()
    log_result("تست عدم ورود بدلیل عدم درج نام کاربری", success=True)

# ========== تست ششم: عدم درج پسورد ==========


@pytest.mark.order(6)
@scenario("../features/login.feature", "Login without password")
def test_Login_without_Password():
    pass


@given("کاربر در صفحه ورود به سامانه قرار دارد")
def open_login_page_without_password_error(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")


@when("کاربر رمز عبور را وارد نمی‌کند")
def fill_without_Password(page):
    sleep(0.5)
    page.get_by_placeholder(" رمز عبور را وارد کنید").fill("")
    
    sleep(0.5)


@when("کاربر نام کاربری و کد امنیتی صحیح را وارد می‌کند")
def fill_correct_username2_and_captcha2(page):
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071000")
    sleep(0.5)
    page.get_by_placeholder("کد امنیتی را وارد کنید").fill("PASSED")
    sleep(0.5)


@when("کاربر روی دکمه ورود کلیک می‌کند")
def click_login_button_without_pass(page):

    sleep(0.5)
    page.get_by_role("button", name="ورود").click()
    sleep(1)


@then(parsers.parse("پیغام خطای ورود با عدم درج پسورد نمایش داده می‌شود"))

@log_step("تست ورود ناموفق با عدم درج پسورد")

def check_without_password_error(page):
    sleep(2)
    expect(page.locator("text='فیلد الزامی است.'")).to_be_visible()
    log_result("تست عدم ورود بدلیل عدم درج پسورد", success=True)


# ========== تست ششم: عدم درج کپچا ==========


@pytest.mark.order(7)
@scenario("../features/login.feature", "Login without captcha")
def test_Login_without_captcha():
    pass


@given("کاربر در صفحه ورود به سامانه قرار دارد")
def open_login_page_without_captcha(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")


@when("کاربر نام کاربری و رمز عبور صحیح را وارد می‌کند")
def fill_valid_username_Password_without_captcha(page):
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071523")
    sleep(0.5)
    page.get_by_placeholder(" رمز عبور را وارد کنید").fill("0081071523")
    
    sleep(0.5)


@when("کاربر کد کپچای را وارد نمی کند")
def fill_without_captcha(page):
    sleep(0.5)
    page.get_by_placeholder("کد امنیتی را وارد کنید").fill("")
    sleep(0.5)


@when("کاربر روی دکمه ورود کلیک می‌کند")
def click_login_button_wrong_user(page):

    sleep(0.5)
    page.get_by_role("button", name="ورود").click()
    sleep(3)


@then(parsers.parse("پیغام خطای ورود با عدم درج کپچا نمایش داده می‌شود"))

@log_step("تست ورود ناموفق با عدم درج کد کپچا")

def check_without_captcha(page):
    sleep(2)
    expect(page.locator("text='فیلد الزامی است.'")).to_be_visible()
    log_result("تست عدم ورود بدلیل عدم درج کپچا", success=True)
