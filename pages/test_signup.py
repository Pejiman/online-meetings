import re
import os
import datetime
import logging
import random
from time import sleep
from pytest_bdd import scenario, when, then, parsers, given
from faker import Faker
from openpyxl import Workbook, load_workbook
from playwright.sync_api import expect
import pytest
from playwright.sync_api import sync_playwright
from functools import wraps

@pytest.fixture(scope="function")
def page():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)  # True برای تست‌های CI
        context = browser.new_context(
            viewport={"width": 1920, "height": 1080},  # نمایشگر بزرگ
            screen={"width": 1920, "height": 1080}
        )
        page = context.new_page()
        yield page
        context.close()
        browser.close()

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

RESULTS_FILE = "test_results.xlsx"
SCREENSHOT_DIR = "./screenshots"
os.makedirs(SCREENSHOT_DIR, exist_ok=True)


def log_result(test_name, success=True, screenshot_path=None):
    date_now = datetime.datetime.now().strftime("%Y-%m-%d")
    time_now = datetime.datetime.now().strftime("%H:%M:%S")
    result = "موفق" if success else "ناموفق"

    if not os.path.exists(RESULTS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["تست", "تاریخ", "ساعت", "نتیجه تست", "مسیر اسکرین‌شات"])
    else:
        wb = load_workbook(RESULTS_FILE)
        ws = wb.active

    ws.append([test_name, date_now, time_now, result, screenshot_path or "-"])
    wb.save(RESULTS_FILE)
    logger.info(f"نتیجه تست '{test_name}' با وضعیت: {result}")


def take_screenshot(page, test_name):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    screenshot_path = os.path.join(SCREENSHOT_DIR, f"{test_name}_{timestamp}.png")
    page.screenshot(path=screenshot_path)
    logger.warning(f"📸 اسکرین‌شات گرفته شد: {screenshot_path}")
    return screenshot_path

def log_step(test_name):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            page = kwargs.get("page", None)
            try:
                return func(*args, **kwargs)
            except Exception:
                if page:
                    screenshot_path = take_screenshot(page, test_name)
                else:
                    screenshot_path = None
                log_result(test_name, success=False, screenshot_path=screenshot_path)
                raise
        return wrapper
    return decorator

def generate_valid_national_code():
    code = [random.randint(0, 9) for _ in range(9)]
    s = sum([(10 - i) * code[i] for i in range(9)])
    r = s % 11
    if r < 2:
        control_digit = r
    else:
        control_digit = 11 - r
    code.append(control_digit)
    return ''.join(map(str, code))


# ========== تست اول: ثبت صحیح ==========

@pytest.mark.order(1)
@scenario("../features/signup.feature", "Correct signup")
def test_signup_Correct():
    pass


@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به صورت صحیح وارد می‌کند")
def fill_signup_Correct_test(page):

    national_code = generate_valid_national_code()

    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("پژمان")
    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").click()
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("رنجی دیزچی")
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill(national_code) 
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("۱۳۶۷/۰۲/۲5")
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)

@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_signup_Correct_test_register(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام موفقیت ثبت‌نام با موفقیت انجام شد نمایش داده می‌شود")
def check_dashboard_loaded(page):
    try:
       
        expect(page.locator("text=ثبت نام با موفقیت انجام شد")).to_be_visible(timeout=15000)
        print("✅ ثبت نام موفق و داشبورد نمایش داده شد.")
        log_result("ثبت‌نام صحیح", success=True)
    except Exception as e:
        print("❌ ثبت نام موفق نبود یا داشبورد ظاهر نشد.")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("ثبت‌نام صحیح", success=True)
        raise e


# ========== تست دوم: ثبت ناموفق با عدم ثبت نام ==========


@pytest.mark.order(2)
@scenario("../features/signup.feature", "signup without name")
def test_signup_without_name():
    pass


@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد نام به صورت صحیح وارد می کند")
def fill_signup_without_name(page):


    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("")
    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").click()
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("رنجی دیزچی")
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071523")  
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("۱۳۶۷/۰۲/۲5")
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)

@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_signup_withoutname(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام خطای وارد کردن نام الزامی است نمایش داده می‌شود")
def check_dashboard_loaded(page):
    try:
       
        expect(page.locator("text=فیلد الزامی است.")).to_be_visible(timeout=15000)
        print("✅ فیلد نام به درستی اجباری می باشد")
        log_result("ثبت‌نام بدون نام", success=True)
    except Exception as e:
        print("❌ فیلد نام به اشتباه اجباری نمی باشد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("ثبت‌نام بدون نام", success=True)
        raise e  

    # ==========  تست سوم: ثبت ناموفق با عدم ثبت نام خانوادگی ==========


@pytest.mark.order(3)
@scenario("../features/signup.feature", "signup without last name")
def test_signup_without_last_name():
    pass


@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد نام خانوادگی به صورت صحیح وارد می کند")
def fill_signup_without_last_name(page):


    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("پژمان")
    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("")
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071523") 
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("۱۳۶۷/۰۲/۲5")
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)

@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_signup_without_last_name(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام خطای وارد کردن نام خانوادگی الزامی است نمایش داده می‌شود")
def check_dashboard_loaded(page):
    try:
       
        expect(page.locator("text=فیلد الزامی است.")).to_be_visible(timeout=15000)
        print("✅ فیلد نام خانوادگی به درستی اجباری می باشد")
        log_result("ثبت‌نام بدون نام خانوادگی", success=True)
    except Exception as e:
        print("❌ فیلد نام خانوادگی به اشتباه اجباری نمی باشد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("ثبت‌نام بدون نام خانوادگی", success=True)
        raise e  


        # ==========  تست چهارم: ثبت ناموفق با عدم ثبت کد ملی ==========


@pytest.mark.order(4)
@scenario("../features/signup.feature", "signup without national code")
def test_signup_without_nationalcode():
    pass


@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد کد ملی به صورت صحیح وارد می کند")
def fill_signup_without_nationalcode(page):


    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("پژمان")
    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("رنجی دیزچی")
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill("") 
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("۱۳۶۷/۰۲/۲5")
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)

@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_signup_without_national_code(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام خطای وارد کردن کد ملی الزامی است نمایش داده می‌شود")
def check_dashboard_loaded(page):
    try:
       
        expect(page.locator("text=کد ملی باید 10 رقم باشد")).to_be_visible(timeout=15000)
        print("✅ فیلد کد ملی به درستی اجباری می باشد")
        log_result("ثبت‌نام بدون کد ملی", success=True)
    except Exception as e:
        print("❌ فیلد کد ملی به اشتباه اجباری نمی باشد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("ثبت‌نام بدون کد ملی", success=True)
        raise e  


 # ==========  تست پنجم: ثبت ناموفق با عدم ثبت موبایل ==========


@pytest.mark.order(5)
@scenario("../features/signup.feature", "signup without mobie.NO")
def test_signup_without_mobie_NO():
    pass


@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد شماره موبایل به صورت صحیح وارد می کند")
def fill_signup_without_mobieNO(page):


    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("پژمان")
    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("رنجی دیزچی")
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071523") 
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("")
    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("۱۳۶۷/۰۲/۲5")
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)

@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_signup_without_mobie_NO(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام خطای وارد کردن شماره موبایل الزامی است نمایش داده می‌شود")
def check_dashboard_loaded(page):
    try:
       
        expect(page.locator("text=شماره موبایل باید 11 رقم باشد")).to_be_visible(timeout=15000)
        print("✅ فیلد شماره موبایل به درستی اجباری می باشد")
        log_result("ثبت‌نام بدون شماره موبایل", success=True)
    except Exception as e:
        print("❌ فیلد شماره موبایل به اشتباه اجباری نمی باشد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("ثبت‌نام بدون شماره موبایل", success=True)
        raise e  



# ==========  تست ششم: ثبت ناموفق با عدم ثبت تاریخ تولد ==========


@pytest.mark.order(6)
@scenario("../features/signup.feature", "signup without date of birth")
def test_signup_without_date_of_birth():
    pass


@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد تاریخ تولد به صورت صحیح وارد می کند")
def fill_signup_without_dateofbirth(page):


    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("پژمان")
    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("رنجی دیزچی")
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071523") 
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)

@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_signup_without_date_of_birth(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام خطای وارد کردن تاریخ تولد الزامی است نمایش داده می‌شود")
def check_dashboard_loaded(page):
    try:
       
        expect(page.locator("text=فیلد الزامی است.")).to_be_visible(timeout=15000)
        print("✅ فیلد تاریخ تولد به درستی اجباری می باشد")
        log_result("ثبت‌نام بدون تاریخ تولد", success=True)
    except Exception as e:
        print("❌ فیلد تاریخ تولد به اشتباه اجباری نمی باشد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("ثبت‌نام بدون تاریخ تولد", success=True)
        raise e  


# ==========  تست هفتم: ثبت ناموفق با عدم ثبت پسورد ==========


@pytest.mark.order(7)
@scenario("../features/signup.feature", "signup without password")
def test_signup_without_password():
    pass


@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد پسورد به صورت صحیح وارد می کند")
def fill_signup_without_password(page):


    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("پژمان")
    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("رنجی دیزچی")
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071523") 
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("۱۳۶۷/۰۲/۲5")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)

@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_signup_withoutpassword(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام خطای وارد کردن پسورد الزامی است نمایش داده می‌شود")
def check_dashboard_loaded(page):
    try:
       
        expect(page.locator("text=فیلد الزامی است.")).to_be_visible(timeout=15000)
        print("✅ فیلد پسورد به درستی اجباری می باشد")
        log_result("ثبت‌نام بدون پسورد", success=True)
    except Exception as e:
        print("❌ فیلد پسورد به اشتباه اجباری نمی باشد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("ثبت‌نام بدون پسورد", success=True)
        raise e  



# ==========  تست هشتم: ثبت ناموفق با عدم ثبت  تکرار پسورد ==========


@pytest.mark.order(8)
@scenario("../features/signup.feature", "signup without re password")
def test_signup_without_re_password():
    pass


@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد تکرار پسورد به صورت صحیح وارد می کند")
def fill_signup_without_repassword(page):


    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("پژمان")
    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("رنجی دیزچی")
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071523") 
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("۱۳۶۷/۰۲/۲5")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("")
    sleep(0.5)

@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_signup_without_re_password(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام خطای وارد کردن تکرار پسورد الزامی است نمایش داده می‌شود")
def check_dashboard_loaded(page):
    try:
       
        expect(page.locator("text=فیلد الزامی است.")).to_be_visible(timeout=15000)
        print("✅ فیلد تکرار پسورد به درستی اجباری می باشد")
        log_result("ثبت‌نام بدون تکرار پسورد", success=True)
    except Exception as e:
        print("❌ فیلد تکرار پسورد به اشتباه اجباری نمی باشد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("ثبت‌نام بدون تکرار پسورد", success=True)
        raise e  



# ==========  تست نهم: ثبت ناموفق با ثبت  کمتر از 2 کاراکتر در فیلد نام ==========

@pytest.mark.order(9)
@scenario("../features/signup.feature", "signup without Enter less than 2 characters in the name field")
def test_signup_Enter_less_than_2_characters_in_the_name_field():
    pass

@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد نام تکمیل می کند")
def fill_signup_Enter_less_than_2_characters_in_the_name_field(page):
    
    national_code2 = generate_valid_national_code()

    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").click()
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("رنجی دیزچی")
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill(national_code2)
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("1367/02/25")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click() 
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)


@when("کاربر فیلد نام را با 1 کاراکتر پر می کند")
def fill_Enter_less_than_2_characters_in_the_name_field(page):

    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("پ")
    sleep(0.5)


@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_Enter_less_than_2charactersinthename_field(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام نام می بایست حداقل 2 کاراکتر باشد نمایش داده می‌شود")
def check_dashboard_loaded_4(page):
    try:
        expect(page.locator("text=نام باید حداقل 2 کاراکتر باشد")).to_be_visible(timeout=15000)
        print("✅ فیلد نام به درستی درصورت درج کاراکتر کمتر از 2 پیغام خطا نمایش می دهد ")
        log_result("درج نام با کمتر از 2 کاراکتر", success=True)
    except Exception as e:
        print("❌ فیلد نام به اشتباه درصورت درج کاراکتر کمتر از 2 پیغام خطا نمایش نمی دهد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("درج نام با کمتر از 2 کاراکتر", success=False)
        raise e


# ==========  تست دهم: ثبت ناموفق با ثبت  بیشتر از 60 کاراکتر در فیلد نام خانوادگی ==========

@pytest.mark.order(10)
@scenario("../features/signup.feature", "signup without Enter more than 60 characters in the last name field")
def test_signup_Enter_more_than_60_characters_in_the_last_name_field():
    pass

@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد نام خانوادگی تکمیل می کند")
def fill_signup_Enter_more_than_60_characters_in_the_last_name_field(page):
    
    national_code3 = generate_valid_national_code()

    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("پژمان")
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill(national_code3)
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("1367/02/25")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click() 
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)


@when("کاربر فیلد نام خانوادگی را با بیش از 60 کاراکتر پر می کند")
def fill_Enter_more_than_60_characters_in_the_last_name_field(page):

    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").click()
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("رنجی دیزجی رنجی دیزجی رنجی دیزجی رنجی دیزجی رنجی دیزجی رنجی دیزجی رنجی دیزجی رنجی دیزجی رنجی دیزجی رنجی دیزجی")
    sleep(0.5)

@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_Enter_more_than_60_characters_in_the_lastname_field(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام نام خانوادگی می بایست حداکثر 60 کاراکتر باشد نمایش داده می‌شود")
def check_dashboard_loaded_7(page):
    try:
        expect(page.locator("text=نام خانوادگی باید حداکثر 60 کاراکتر باشد")).to_be_visible(timeout=15000)
        print("✅ فیلد نام خانوادگی به درستی درصورت درج کاراکتر بیشتر از 60 پیغام خطا نمایش می دهد ")
        log_result("درج نام خانوادکی با بیش از 60 کاراکتر", success=True)
    except Exception as e:
        print("❌ فیلد نام خانوادگی به اشتباه درصورت درج کاراکتر بیشتر از 60 پیغام خطا نمایش نمی دهد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("درج نام خانوادکی با بیش از 60 کاراکتر", success=False)
        raise e



# ==========  تست یازدهم: ثبت ناموفق با ثبت  کد ملی کمتر از 10 کاراکتر ==========


@pytest.mark.order(11)
@scenario("../features/signup.feature", "signup without national code less than 10 characters")
def test_signup_without_nationalcodeless():
    pass


@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد کد ملی تکمیل می کند")
def fill_signup_without_nationalcodeless(page):


    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("پژمان")
    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("رنجی دیزچی")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("۱۳۶۷/۰۲/۲5")
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)


@when("کاربر فیلد کد ملی را با کمتر از 10 کاراکتر پر می کند")
def fill_national_code_less_than_10_character(page):

    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill("123456")    
    sleep(0.5)


@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_signup_without_national_code_less(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام کد ملی باید 10 رقم باشد نمایش داده می‌شود")
def check_dashboard_loaded(page):
    try:
       
        expect(page.locator("text=کد ملی باید 10 رقم باشد")).to_be_visible(timeout=15000)
        print("✅ فیلد کد ملی به درستی 10 کاراکتر می باشد")
        log_result("ثبت‌نام با کد ملی کمتر از 10 کاراکتر", success=True)
    except Exception as e:
        print("❌ فیلد کد ملی به اشتباه 10 کاراکتر نمی باشد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("ثبت‌نام با کد ملی کمتر از 10 کاراکتر", success=True)
        raise e  


# ==========  تست دوازدهم: ثبت ناموفق با ثبت  کد ملی نا معتبر ==========

@pytest.mark.order(12)
@scenario("../features/signup.feature", "signup without Invalid national code")
def test_signup_without_Invalid_national_code():
    pass


@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد کد ملی تکمیل می کند")
def fill_signup_without_Invalidnationalcode(page):


    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("پژمان")
    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("رنجی دیزچی")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("۱۳۶۷/۰۲/۲5")
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)


@when("کاربر فیلد کد ملی را با کد ملی نا معتبر پر می کند")
def fill_Invalid_national_code(page):

    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill("1234567890")    
    sleep(0.5)


@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_signup_without_Invalid_national_code(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام کد ملی صحیح نمی باشد نمایش داده می‌شود")
def check_dashboard_loaded(page):
    try:
       
        expect(page.locator("text=کد ملی وارد شده اشتباه است")).to_be_visible(timeout=15000)
        print("✅ خطای کد ملی نا معتبر به درستی نمایش داده شده")
        log_result("ثبت‌نام با کد ملی نا معتبر", success=True)
    except Exception as e:
        print("❌ خطای کد ملی نا معتبر به اشتباه نمایش داده نشده")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("ثبت‌نام با کد ملی نا معتبر", success=True)
        raise e  



# ==========  تست سیزدهم: ثبت ناموفق با ثبت شماره موبایل کمتر از 11 کاراکتر ==========


@pytest.mark.order(13)
@scenario("../features/signup.feature", "signup without mobie.NO less than 11 characters")
def test_signup_without_mobie_NO_less_than_characters():
    pass


@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد شماره موبایل تکمیل می کند")
def fill_signup_without_mobieNO_lessthan_characters(page):


    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("پژمان")
    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("رنجی دیزچی")
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071523")
    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("۱۳۶۷/۰۲/۲5")
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)


@when("کاربر فیلد شماره موبایل را با کمتر از 11 کاراکتر پر می کند")
def fill__mobie_NO_less_than_characters(page):

    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("091252436")    
    sleep(0.5)


@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_signup_without_mobieNO_lessthan_characters(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام شماره موبایل باید 11 رقم باشد نمایش داده می‌شود")
def check_dashboard_loaded(page):
    try:
       
        expect(page.locator("text=شماره موبایل باید 11 رقم باشد")).to_be_visible(timeout=15000)
        print("✅ فیلد شماره موبایل به درستی 11 کاراکتر می باشد")
        log_result("ثبت‌نام با شماره موبایل کمتر از 11 کاراکتر", success=True)
    except Exception as e:
        print("❌ فیلد شماره موبایل به اشتباه 11 کاراکتر نمی باشد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("ثبت‌نام با شماره موبایل کمتر از 11 کاراکتر", success=True)
        raise e  


    
# ==========  تست چهاردهم: ثبت ناموفق با ثبت تاریخ تولد کمتر از 18 سال

@pytest.mark.order(14)
@scenario("../features/signup.feature", "signup without Minimum age requirement")
def test_signup_without_Minimum_age_requirement():
    pass


@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد تاریخ تولد تکمیل می کند")
def fill_signup_without_Minimum_age_requirement(page):


    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("پژمان")
    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("رنجی دیزچی")
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071523")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)


@when("کاربر فیلد تاریخ تولد را برای سن کمتر از 18 سال پر می کند")
def fill__Minimum_age_requirement(page):

    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("1400/02/25")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click() 
    sleep(0.5)


@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_signup_without_Minimumage_requirement(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام تاریخ تولد باید بزرگتر از 18 سال باشد نمایش داده می‌شود")
def check_dashboard_loaded(page):
    try:
       
        expect(page.locator("text=شما باید بیش از 18 ساله و کمتر از 120 ساله باشید")).to_be_visible(timeout=15000)
        print("✅ فیلد تاریخ تولد به درستی بیش از 18 ساله و کمتر از 120 ساله می باشد")
        log_result("ثبت‌نام با تاریخ تولد بیشتر از 18 سال", success=True)
    except Exception as e:
        print("❌ فیلد شماره موبایل به اشتباه بیش از 18 ساله و کمتر از 120 ساله نمی باشد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("ثبت‌نام با تاریخ تولد بیشتر از 18 سال", success=True)
        raise e  


      
# ==========  تست پانزده: ثبت ناموفق با ثبت تاریخ تولد بیشتر از 120 سال

@pytest.mark.order(15)
@scenario("../features/signup.feature", "signup without Maximum age requirement")
def test_signup_without_Maximum_age_requirement():
    pass


@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد تاریخ تولد تکمیل می کند")
def fill_signup_without_Maximum_age_requirement(page):


    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("پژمان")
    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("رنجی دیزچی")
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071523")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)


@when("کاربر فیلد تاریخ تولد را برای سن بیشتر از 120 سال پر می کند")
def fill__Maximum_age_requirement(page):

    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("1200/02/25")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click() 
    sleep(0.5)


@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_signup_without_Maximumage_requirement(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام تاریخ تولد باید کوچکتر از 120 سال باشد نمایش داده می‌شود")
def check_dashboard_loaded(page):
    try:
       
        expect(page.locator("text=شما باید بیش از 18 ساله و کمتر از 120 ساله باشید")).to_be_visible(timeout=15000)
        print("✅ فیلد تاریخ تولد به درستی بیش از 18 ساله و کمتر از 120 ساله می باشد")
        log_result("ثبت‌نام با تاریخ تولد کمتر از 120 سال", success=True)
    except Exception as e:
        print("❌ فیلد شماره موبایل به اشتباه بیش از 18 ساله و کمتر از 120 ساله نمی باشد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("ثبت‌نام با تاریخ تولد کمتر از 120 سال", success=True)
        raise e  





    # ==========  تست شانزدهم: ثبت ناموفق با ثبت پسورد ساده

@pytest.mark.order(16)
@scenario("../features/signup.feature", "signup without easy password")
def test_signup_without_easy_password():
    pass


@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد پسورد تکمیل می کند")
def fill_signup_without_easypassword(page):


    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("پژمان")
    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("رنجی دیزچی")
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071523")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("1367/02/25")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click() 
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)


@when("کاربر فیلد پسورد را با تعداد کاراکتر کمتر از 8 کاراکتر تکمیل می کند")
def fill__easy_password_11(page):

    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("123456")
    sleep(0.5)
   


@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_signup_without_easy_password(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام تعداد کاراکتر می بایست حداقل 8 کاراکتر باشد نمایش داده می‌شود")
def check_dashboard_loaded_1(page):
    try:
       
        expect(page.locator("text=رمز عبور باید حداقل 8 رقم باشد")).to_be_visible(timeout=15000)
        print("✅ فیلد رمز عبور به درستی نمیتواند کمتر از 8 کاراکتر باشد")
        log_result("رمز عبور با کمتر از 8 کاراکتر", success=True)
    except Exception as e:
        print("❌ فیلد رمز عبور به اشتباه میتواند کمتر از 8 کاراکتر باشد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("رمز عبور با کمتر از 8 کاراکتر", success=True)
        raise e  



@when("کاربر فیلد پسورد را با بدون وارد کردن حروف بزرگ تکمیل می کند")
def fill__easy_password_1(page):

    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("12345678abs")
    sleep(0.5)
    page.get_by_role("button", name="ثبت نام").click()
    sleep(0.5)



@then("پیام رمز عبور باید حداقل یک حرف بزرگ داشته باشد نمایش داده می‌شود")
def check_dashboard_loaded_2(page):
    try:
       
        expect(page.locator("text=رمز عبور باید حداقل یک حرف بزرگ داشته باشد")).to_be_visible(timeout=15000)
        print("✅ فیلد رمز عبور به درستی نمیتواند حداقل یک حرف بزرگ داشته باشد")
        log_result("رمز عبور با حداقل یک حرف بزرگ", success=True)
    except Exception as e:
        print("❌ فیلد رمز عبور به اشتباه میتواند یک حرف بزرگ نداشته باشد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("رمز عبور با حداقل یک حرف بزرگ", success=True)
        raise e  



@when("کاربر فیلد پسورد را بدون وارد کردن علامت خاص تکمیل می کند")
def fill__easy_password_2(page):

    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("12345678abS")
    sleep(0.5)
    page.get_by_role("button", name="ثبت نام").click()
    sleep(0.5)
   

@then("پیام رمز عبور باید حداقل یک حرف خاص داشته باشد نمایش داده می‌شود")
def check_dashboard_loaded_3(page):
    try:
       
        expect(page.locator("text=رمز عبور باید حداقل یک علامت خاص داشته باشد")).to_be_visible(timeout=15000)
        print("✅ فیلد رمز عبور به درستی نمیتواند حداقل یک حرف خاص داشته باشد")
        log_result("رمز عبور با حداقل یک حرف خاص", success=True)
    except Exception as e:
        print("❌ فیلد رمز عبور به اشتباه میتواند یک حرف خاص نداشته باشد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("رمز عبور با حداقل یک حرف خاص", success=True)
        raise e  



    
    # ==========  تست هفدهم: ثبت ناموفق با ثبت  تکرار رمز عبور متفاوت از رمز عبور
@pytest.mark.order(17)
@scenario("../features/signup.feature", "signup without repeat incorrect password")
def test_signup_without_repeat_incorrect_password():
    pass


@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد تکرار رمز عبور تکمیل می کند")
def fill_signup_without_repeatincorrectpassword(page):


    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("پژمان")
    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("رنجی دیزچی")
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071523")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("1367/02/25")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click() 
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)


@when("کاربر فیلد تکرار رمز عبور را برای متفاوت ار فیلد رمز عبور پر می کند")
def fill__repeat_incorrect_password(page):

    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662617")
    sleep(0.5)
   


@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_signup_without_repeat_incorrect_password(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام رمز عبور و تکرار آن یکسان نمی باشد نمایش داده می‌شود")
def check_dashboard_loaded(page):
    try:
       
        expect(page.locator("text=رمز عبور و تکرار آن یکسان نمی باشد")).to_be_visible(timeout=15000)
        print("✅ فیلد تکرار رمز عبور به درستی نمیتواند متفاوت از فیلد رمز عبور باشد")
        log_result("تکرار رمز عبور نا برابر با رمز عبور", success=True)
    except Exception as e:
        print("❌ فیلد تکرار رمز عبور به اشتباه میتواند متفاوت از فیلد رمز عبور باشد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("تکرار رمز عبور نا برابر با رمز عبور", success=True)
        raise e  



    # ==========  تست هجدهم: ثبت ناموفق با ثبت  نام و نام خانوادگی با کاراکتر نا مجاز

@pytest.mark.order(18)
@scenario("../features/signup.feature", "signup without First and last name illegal characters")
def test_signup_without_First_and_last_name_illegal_characters():
    pass

@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد نام و نام خانوادگی تکمیل می کند")
def fill_signup_without_First_last_name_illegal_characters(page):
    
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071523")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("1367/02/25")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click() 
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)


@when("کاربر فیلد نام را با کاراکتر های نا مجاز پر می کند")
def fill__First_name_illegal_characters(page):

    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("12@#$%")
    sleep(0.5)
   


@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_signup_without_First_and_illegal_characters(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام نام باید به صورت فارسی باشد نمایش داده می‌شود")
def check_dashboard_loaded_1(page):
    try:
       
        expect(page.locator("text=نام باید به صورت فارسی باشد")).to_be_visible(timeout=15000)
        print("✅ فیلد نام به درستی درصورت درج کاراکتر نا معتبر پیغام خطا نمایش می دهد ")
        log_result("درج نام با کاراکتر های نا معتبر", success=True)
    except Exception as e:
        print("❌ فیلد نام به اشتباه درصورت درج کاراکتر نا معتبر پیغام خطا نمایش نمی دهد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("درج نام با کاراکتر های نا معتبر", success=True)
        raise e  



@when("کاربر فیلد نام خانوادگی را با کاراکتر های نا مجاز پر می کند")
def fill__last_name_illegal_characters(page):

    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").click()
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("12@#$%")
    sleep(0.5)
    page.get_by_role("button", name="ثبت نام").click()
    sleep(0.5)



@then("پیام نام خانوادگی باید به صورت فارسی باشد نمایش داده می‌شود")
def check_dashboard_loaded_2(page):
    try:
       
        expect(page.locator("text=نام خانوادگی باید به صورت فارسی باشد")).to_be_visible(timeout=15000)
        print("✅ فیلد نام خانوادگی به درستی درصورت درج کاراکتر نا معتبر پیغام خطا نمایش می دهد ")
        log_result("درج نام خانوادگی با کاراکتر های نا معتبر", success=True)
    except Exception as e:
        print("❌ فیلد نام خانوادگی به اشتباه درصورت درج کاراکتر نا معتبر پیغام خطا نمایش نمی دهد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("درج نام خانوادگی با کاراکتر های نا معتبر", success=True)
        raise e  



    # ==========  تست نوزدهم: ثبت ناموفق با ثبت  نام و نام خانوادگی با کاراکتر لاتین

@pytest.mark.order(19)
@scenario("../features/signup.feature", "signup without English first and last name")
def test_signup_without_English_first_and_last_name():
    pass

@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد نام و نام خانوادگی تکمیل می کند")
def fill_signup_without_English_first_last_name(page):
    
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071523")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("1367/02/25")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click() 
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)


@when("کاربر فیلد نام را با کاراکتر های انگلیسی پر می کند")
def fill_English_first_name(page):

    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("Pejman")
    sleep(0.5)
   


@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_signup_without_English_first_name(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام نام باید به صورت فارسی باشد نمایش داده می‌شود")
def check_dashboard_loaded_3(page):
    try:
       
        expect(page.locator("text=نام باید به صورت فارسی باشد")).to_be_visible(timeout=15000)
        print("✅ فیلد نام به درستی درصورت درج کاراکتر انگلیسی پیغام خطا نمایش می دهد ")
        log_result("درج نام با کاراکتر های انگلیسی", success=True)
    except Exception as e:
        print("❌ فیلد نام به اشتباه درصورت درج کاراکتر انگلیسی پیغام خطا نمایش نمی دهد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("درج نام با کاراکتر های انگلیسی", success=True)
        raise e  



@when("کاربر فیلد نام خانوادگی را با کاراکتر های انگلیسی پر می کند")
def fill_English_last_name(page):

    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").click()
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("Ranji")
    sleep(0.5)
    page.get_by_role("button", name="ثبت نام").click()
    sleep(0.5)



@then("پیام نام خانوادگی باید به صورت فارسی باشد نمایش داده می‌شود")
def check_dashboard_loaded_4(page):
    try:
       
        expect(page.locator("text=نام خانوادگی باید به صورت فارسی باشد")).to_be_visible(timeout=15000)
        print("✅ فیلد نام خانوادگی به درستی درصورت درج کاراکتر انگلیسی پیغام خطا نمایش می دهد ")
        log_result("درج نام خانوادگی با کاراکتر های انگلیسی", success=True)
    except Exception as e:
        print("❌ فیلد نام خانوادگی به اشتباه درصورت درج کاراکتر انگلیسی پیغام خطا نمایش نمی دهد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("درج نام خانوادگی با کاراکتر های انگلیسی", success=True)
        raise e  



# ==========  تست بیستم: ثبت ناموفق با ثبت  کد ملی تکراری ==========

@pytest.mark.order(20)
@scenario("../features/signup.feature", "signup without Duplicate national code")
def test_signup_without_Duplicate_national_code():
    pass


@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به جز فیلد کد ملی تکمیل می کند")
def fill_signup_without_Duplicate_nationalcode(page):


    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("پژمان")
    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("رنجی دیزچی")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("۱۳۶۷/۰۲/۲5")
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)


@when("کاربر فیلد کد ملی را با کد ملی تکراری پر می کند")
def fill_Duplicate_national_code(page):

    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill("0081071523")    
    sleep(0.5)


@when("کاربر روی دکمه ثبت‌نام کلیک می‌کند")
def fill_signup_without_Duplicate_national_code(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)

@then("پیام متقاضی با این کد ملی قبلا ثبت شده است نمایش داده می‌شود")
def check_dashboard_loaded(page):
    try:
       
        expect(page.locator("text=متقاضی با این کد ملی قبلا ثبت شده است.")).to_be_visible(timeout=15000)
        print("✅ خطای کد ملی تکراری به درستی نمایش داده شده")
        log_result("ثبت‌نام با کد ملی تکراری", success=True)
    except Exception as e:
        print("❌ خطای کد ملی تکراری به اشتباه نمایش داده نشده")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("ثبت‌نام با کد ملی تکراری", success=True)
        raise e  
 

      
# # ==========  تست بیست و یکم: ثبت ناموفق با ثبت تاریخ تولد با فرمت نادرست

@pytest.mark.order(21)
@scenario("../features/signup.feature", "signup without Birth date in wrong format")
def test_signup_without_Birth_date_in_wrongformat():
    pass


@given("کاربر در صفحه ثبت نام قرار دارد")
def open_login_page1(page):
    page.goto("https://online-meetings-test.rayanbourse.ir/auth/login/")
    sleep(0.5)
    page.get_by_role("link", name="ثبت نام کنید").click()
    sleep(0.5)

@when("کاربر تمام فیلد ها را به غیر از فیلد تاریخ تولد تکمیل می کند")
def fill_signup_without_Birth_date_in_wrongformat1(page):

    national_code1 = generate_valid_national_code()
    page.get_by_placeholder("نام را وارد کنید").click()
    sleep(0.5)
    page.get_by_placeholder("نام را وارد کنید").fill("پژمان")
    sleep(0.5)
    page.get_by_placeholder("نام خانوادگی را وارد کنید").fill("رنجی دیزچی")
    sleep(0.5)
    page.get_by_placeholder("کد ملی را وارد کنید").click()
    page.get_by_placeholder("کد ملی را وارد کنید").fill(national_code1)
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click()
    page.get_by_placeholder("شماره موبایل را وارد کنید").fill("09125243681")
    sleep(0.5)
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).click()
    page.get_by_placeholder("رمز عبور را وارد کنید", exact=True).fill("Pejm@n44662618")
    sleep(0.5)
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").click()
    page.get_by_placeholder("تکرار رمز عبور را وارد کنید").fill("Pejm@n44662618")
    sleep(0.5)


@when("کاربر فیلد تاریخ تولد را با داده های نا معتبر پر می کند")
def fill_Birth_date_in_wrongformat11(page):

    sleep(0.5)
    page.get_by_placeholder("تاریخ تولد").click()
    page.get_by_placeholder("تاریخ تولد").fill("1367+02+25")
    sleep(0.5)
    page.get_by_placeholder("شماره موبایل را وارد کنید").click() 
    sleep(0.5)


@when("کاربر روی دکمه ثبت‌ نام کلیک می‌کند")
def fill_signup_without_wrongformat11(page):

   page.get_by_role("button", name="ثبت نام").click()
   sleep(0.5)


@then("پیام فرمت تاریخ تولد صحیح نمی باشد نمایش داده می‌شود")
def check_dashboard_loaded(page):
    try:
        expect(page.locator("text=فرمت تاریخ تولد درست نمی باشد")).to_be_visible(timeout=15000)
        print("✅ فرمت فیلد تاریخ تولد به درستی خطا گرفته شد")
        log_result("ثبت‌نام با تاریخ تولد با فرمت اشتباه", success=True)
    except Exception as e:
        print("❌ فرمت فیلد تاریخ تولد به اشتباه خطا گرفته نشد")
        page.screenshot(path="./screenshots/failure_dashboard.png")
        log_result("ثبت‌نام با تاریخ تولد با فرمت اشتباه", success=False)  
        raise e
